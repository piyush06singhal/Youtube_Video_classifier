
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from googleapiclient.discovery import build
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import anvil.server
import anvil.media
import re
import emoji
from io import BytesIO
from datetime import datetime

# Initialize VADER sentiment analyzer (much better than TextBlob for social media)
vader_analyzer = SentimentIntensityAnalyzer()

# Connect to Anvil
anvil.server.connect("server_RM5P54PWIXQ3JIVF2AGIY6O4-PZQUJ5L7AZPWN6JK")

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def clean_comment_text(text):
    """Clean comment text by removing HTML tags, extra spaces, and URLs"""
    # Remove HTML tags
    text = re.sub(r'<[^>]+>', '', text)
    # Remove URLs
    text = re.sub(r'http\S+|www.\S+', '', text)
    # Remove extra whitespace
    text = ' '.join(text.split())
    return text

def convert_emojis_to_text(text):
    """Convert emojis to text descriptions for better sentiment analysis"""
    return emoji.demojize(text, delimiters=(" ", " "))

def detect_spam(comment):
    """Simple spam detection based on patterns"""
    spam_indicators = [
        len(comment) < 5,  # Too short
        comment.count('http') > 2,  # Too many links
        comment.isupper() and len(comment) > 20,  # All caps and long
        comment.count('!') > 5,  # Too many exclamation marks
        len(set(comment.split())) < len(comment.split()) * 0.3  # Too repetitive
    ]
    return sum(spam_indicators) >= 2

def detect_bot(username, comment):
    """Simple bot detection based on username and comment patterns"""
    bot_indicators = [
        'bot' in username.lower(),
        username.endswith(tuple('0123456789' * 3)),  # Ends with many numbers
        len(comment.split()) < 3,  # Very short comment
        comment.count('@') > 2  # Too many mentions
    ]
    return sum(bot_indicators) >= 2

# ============================================================================
# YOUTUBE API FUNCTIONS
# ============================================================================

def fetch_video_metadata(video_id, api_key):
    """Fetch video metadata like title, views, channel info"""
    try:
        youtube_service = build("youtube", "v3", developerKey=api_key)
        request = youtube_service.videos().list(
            part="snippet,statistics",
            id=video_id
        )
        response = request.execute()
        
        if not response.get('items'):
            return None
            
        video_data = response['items'][0]
        metadata = {
            'title': video_data['snippet']['title'],
            'channel': video_data['snippet']['channelTitle'],
            'published_at': video_data['snippet']['publishedAt'],
            'view_count': video_data['statistics'].get('viewCount', 'N/A'),
            'like_count': video_data['statistics'].get('likeCount', 'N/A'),
            'comment_count': video_data['statistics'].get('commentCount', 'N/A')
        }
        return metadata
    except Exception as e:
        print(f"Error fetching metadata: {e}")
        return None

def fetch_comments(video_id, api_key, max_comments=1000):
    """Fetch comments from YouTube with improved error handling"""
    youtube_service = build("youtube", "v3", developerKey=api_key)
    all_comments = []
    
    try:
        api_request = youtube_service.commentThreads().list(
            part="snippet",
            videoId=video_id,
            maxResults=100,
            textFormat="plainText"
        )
        
        while api_request is not None and len(all_comments) < max_comments:
            api_response = api_request.execute()
            
            for comment_data in api_response.get("items", []):
                snippet = comment_data["snippet"]["topLevelComment"]["snippet"]
                username = snippet["authorDisplayName"]
                text = snippet["textDisplay"]
                published_at = snippet.get("publishedAt", "")
                like_count = snippet.get("likeCount", 0)
                
                all_comments.append({
                    'username': username,
                    'text': text,
                    'published_at': published_at,
                    'like_count': like_count
                })
            
            # Check for next page
            next_page_token = api_response.get("nextPageToken")
            if next_page_token and len(all_comments) < max_comments:
                api_request = youtube_service.commentThreads().list(
                    part="snippet",
                    videoId=video_id,
                    maxResults=100,
                    pageToken=next_page_token,
                    textFormat="plainText"
                )
            else:
                break
                
    except Exception as e:
        print(f"Error fetching comments: {e}")
        if "commentsDisabled" in str(e):
            raise Exception("Comments are disabled for this video")
        raise e
    
    return all_comments

# ============================================================================
# SENTIMENT ANALYSIS FUNCTIONS
# ============================================================================

def analyze_comment_sentiment(comment_text):
    """Analyze sentiment of a single comment using VADER"""
    # Clean and prepare text
    cleaned_text = clean_comment_text(comment_text)
    emoji_text = convert_emojis_to_text(cleaned_text)
    
    # Get VADER scores
    scores = vader_analyzer.polarity_scores(emoji_text)
    compound_score = scores['compound']
    
    # Classify sentiment based on compound score
    if compound_score >= 0.5:
        sentiment_label = "Extremely Positive"
    elif compound_score >= 0.05:
        sentiment_label = "Positive"
    elif compound_score > -0.05:
        sentiment_label = "Neutral"
    elif compound_score > -0.5:
        sentiment_label = "Negative"
    else:
        sentiment_label = "Extremely Negative"
    
    return {
        'score': round(compound_score, 4),
        'label': sentiment_label,
        'positive': round(scores['pos'], 4),
        'neutral': round(scores['neu'], 4),
        'negative': round(scores['neg'], 4)
    }

def analyze_all_comments(comments):
    """Analyze sentiment for all comments"""
    analyzed_comments = []
    
    for comment in comments:
        sentiment = analyze_comment_sentiment(comment['text'])
        is_spam = detect_spam(comment['text'])
        is_bot = detect_bot(comment['username'], comment['text'])
        
        analyzed_comments.append({
            'username': comment['username'],
            'text': comment['text'],
            'sentiment_score': sentiment['score'],
            'sentiment_label': sentiment['label'],
            'positive_score': sentiment['positive'],
            'neutral_score': sentiment['neutral'],
            'negative_score': sentiment['negative'],
            'like_count': comment.get('like_count', 0),
            'is_spam': is_spam,
            'is_bot': is_bot,
            'published_at': comment.get('published_at', '')
        })
    
    return analyzed_comments

def calculate_overall_sentiment(analyzed_comments):
    """Calculate overall sentiment from all comments"""
    # Filter out spam and bot comments for overall sentiment
    valid_comments = [c for c in analyzed_comments if not c['is_spam'] and not c['is_bot']]
    
    if not valid_comments:
        return "Neutral", 0.0
    
    # Calculate average sentiment score
    avg_score = sum(c['sentiment_score'] for c in valid_comments) / len(valid_comments)
    
    # Classify overall sentiment
    if avg_score >= 0.5:
        overall_label = "Extremely Positive"
    elif avg_score >= 0.05:
        overall_label = "Positive"
    elif avg_score > -0.05:
        overall_label = "Neutral"
    elif avg_score > -0.5:
        overall_label = "Negative"
    else:
        overall_label = "Extremely Negative"
    
    return overall_label, round(avg_score, 4)

# ============================================================================
# GENRE CLASSIFICATION FUNCTIONS
# ============================================================================

def classify_genre(analyzed_comments):
    """Enhanced genre classification with better keyword matching"""
    genres = {
        "Comedy": ["funny", "lol", "lmao", "hilarious", "laugh", "comedy", "humor", "haha", "😂", "🤣"],
        "Educational": ["learn", "educational", "informative", "tutorial", "explained", "knowledge", "teach"],
        "Entertainment": ["entertaining", "fun", "amazing", "awesome", "cool", "great", "love it"],
        "Music": ["song", "music", "beat", "lyrics", "singer", "album", "track", "melody"],
        "Gaming": ["game", "gaming", "play", "player", "level", "gameplay", "gamer"],
        "Tech": ["technology", "tech", "software", "hardware", "coding", "programming", "app"],
        "Sports": ["sport", "game", "match", "player", "team", "win", "goal", "score"],
        "Food": ["food", "recipe", "cooking", "delicious", "taste", "chef", "restaurant"],
        "Travel": ["travel", "trip", "journey", "adventure", "explore", "destination", "vacation"],
        "Fashion": ["fashion", "style", "outfit", "clothes", "dress", "trend", "beauty"],
        "Horror": ["scary", "horror", "terrifying", "frightening", "creepy", "spooky", "nightmare"],
        "Action": ["action", "fight", "battle", "explosion", "intense", "epic"],
        "Romance": ["romantic", "love", "relationship", "couple", "heart", "❤️", "💕"],
        "Motivational": ["inspiring", "motivational", "motivation", "inspire", "success", "goals"],
        "Relaxing": ["relaxing", "calm", "peaceful", "soothing", "chill", "meditation"],
        "Drama": ["drama", "emotional", "touching", "sad", "cry", "tears"],
        "Documentary": ["documentary", "real", "facts", "history", "true story"],
        "Vlog": ["vlog", "daily", "life", "day in", "routine", "behind the scenes"]
    }
    
    genre_scores = {genre: 0 for genre in genres.keys()}
    
    # Count keyword matches in comments
    for comment in analyzed_comments:
        text_lower = comment['text'].lower()
        for genre, keywords in genres.items():
            for keyword in keywords:
                if keyword in text_lower:
                    genre_scores[genre] += 1
    
    # Get top 3 genres
    sorted_genres = sorted(genre_scores.items(), key=lambda x: x[1], reverse=True)
    top_genres = [genre for genre, count in sorted_genres[:3] if count > 0]
    
    if not top_genres:
        return "General"
    elif len(top_genres) == 1:
        return top_genres[0]
    else:
        return ", ".join(top_genres)

# ============================================================================
# EXCEL GENERATION FUNCTIONS
# ============================================================================

def create_excel_file(analyzed_comments, video_metadata, overall_sentiment, overall_score, genre):
    """Create Excel file with all analysis data"""
    wb = openpyxl.Workbook()
    
    # ===== SUMMARY SHEET =====
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Add summary information
    ws_summary['A1'] = "YouTube Video Sentiment Analysis Report"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')
    
    row = 3
    if video_metadata:
        ws_summary[f'A{row}'] = "Video Title:"
        ws_summary[f'B{row}'] = video_metadata['title']
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws_summary[f'A{row}'] = "Channel:"
        ws_summary[f'B{row}'] = video_metadata['channel']
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws_summary[f'A{row}'] = "Views:"
        ws_summary[f'B{row}'] = video_metadata['view_count']
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws_summary[f'A{row}'] = "Likes:"
        ws_summary[f'B{row}'] = video_metadata['like_count']
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 2
    
    ws_summary[f'A{row}'] = "Overall Sentiment:"
    ws_summary[f'B{row}'] = overall_sentiment
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'].font = Font(bold=True, size=12, color="FF0000" if "Negative" in overall_sentiment else "00B050")
    row += 1
    
    ws_summary[f'A{row}'] = "Sentiment Score:"
    ws_summary[f'B{row}'] = overall_score
    ws_summary[f'A{row}'].font = Font(bold=True)
    row += 1
    
    ws_summary[f'A{row}'] = "Genre/Category:"
    ws_summary[f'B{row}'] = genre
    ws_summary[f'A{row}'].font = Font(bold=True)
    row += 1
    
    ws_summary[f'A{row}'] = "Total Comments Analyzed:"
    ws_summary[f'B{row}'] = len(analyzed_comments)
    ws_summary[f'A{row}'].font = Font(bold=True)
    row += 2
    
    # Sentiment distribution
    sentiment_counts = {}
    for comment in analyzed_comments:
        label = comment['sentiment_label']
        sentiment_counts[label] = sentiment_counts.get(label, 0) + 1
    
    ws_summary[f'A{row}'] = "Sentiment Distribution:"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    for sentiment_label, count in sentiment_counts.items():
        percentage = (count / len(analyzed_comments)) * 100
        ws_summary[f'A{row}'] = sentiment_label
        ws_summary[f'B{row}'] = count
        ws_summary[f'C{row}'] = f"{percentage:.1f}%"
        row += 1
    
    # ===== DETAILED COMMENTS SHEET =====
    ws_comments = wb.create_sheet("Detailed Comments")
    
    # Headers
    headers = ["#", "Username", "Comment", "Sentiment", "Score", "Positive", "Neutral", "Negative", "Likes", "Spam?", "Bot?"]
    for col, header in enumerate(headers, 1):
        cell = ws_comments.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add comment data
    for idx, comment in enumerate(analyzed_comments, 1):
        ws_comments.cell(idx + 1, 1, idx)
        ws_comments.cell(idx + 1, 2, comment['username'])
        ws_comments.cell(idx + 1, 3, comment['text'][:500])  # Limit text length
        ws_comments.cell(idx + 1, 4, comment['sentiment_label'])
        ws_comments.cell(idx + 1, 5, comment['sentiment_score'])
        ws_comments.cell(idx + 1, 6, comment['positive_score'])
        ws_comments.cell(idx + 1, 7, comment['neutral_score'])
        ws_comments.cell(idx + 1, 8, comment['negative_score'])
        ws_comments.cell(idx + 1, 9, comment['like_count'])
        ws_comments.cell(idx + 1, 10, "Yes" if comment['is_spam'] else "No")
        ws_comments.cell(idx + 1, 11, "Yes" if comment['is_bot'] else "No")
    
    # Adjust column widths
    ws_comments.column_dimensions['A'].width = 5
    ws_comments.column_dimensions['B'].width = 20
    ws_comments.column_dimensions['C'].width = 60
    ws_comments.column_dimensions['D'].width = 18
    ws_comments.column_dimensions['E'].width = 10
    ws_comments.column_dimensions['F'].width = 10
    ws_comments.column_dimensions['G'].width = 10
    ws_comments.column_dimensions['H'].width = 10
    ws_comments.column_dimensions['I'].width = 8
    ws_comments.column_dimensions['J'].width = 8
    ws_comments.column_dimensions['K'].width = 8
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

# ============================================================================
# ANVIL CALLABLE FUNCTIONS
# ============================================================================

@anvil.server.callable("analyze_video")
def analyze_video(video_url):
    """Main function to analyze video and return results"""
    api_key = ""
    
    # Extract video ID
    video_id = None
    if "v=" in video_url:
        video_id = re.search(r"(?<=v=)[^&]+", video_url).group(0)
    elif "youtu.be/" in video_url:
        video_id = re.search(r"youtu\.be/([^?]+)", video_url).group(1)
    
    if not video_id:
        return {"error": "Invalid YouTube URL. Please provide a valid YouTube video link."}
    
    try:
        # Fetch video metadata
        print("Fetching video metadata...")
        video_metadata = fetch_video_metadata(video_id, api_key)
        
        # Fetch comments
        print("Fetching comments...")
        comments = fetch_comments(video_id, api_key, max_comments=1000)
        
        if not comments:
            return {"error": "No comments found for this video. Comments might be disabled."}
        
        print(f"Analyzing {len(comments)} comments...")
        
        # Analyze all comments
        analyzed_comments = analyze_all_comments(comments)
        
        # Calculate overall sentiment
        overall_sentiment, overall_score = calculate_overall_sentiment(analyzed_comments)
        
        # Classify genre
        genre = classify_genre(analyzed_comments)
        
        # Calculate statistics
        total_comments = len(analyzed_comments)
        spam_count = sum(1 for c in analyzed_comments if c['is_spam'])
        bot_count = sum(1 for c in analyzed_comments if c['is_bot'])
        
        print("Analysis complete!")
        
        return {
            "sentiment": overall_sentiment,
            "sentiment_score": overall_score,
            "genre": genre,
            "total_comments": total_comments,
            "spam_count": spam_count,
            "bot_count": bot_count,
            "video_title": video_metadata['title'] if video_metadata else "Unknown",
            "video_channel": video_metadata['channel'] if video_metadata else "Unknown",
            "message": "Analysis complete! Click 'Download Report' to get the detailed Excel file."
        }
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return {"error": f"An error occurred: {str(e)}"}

@anvil.server.callable("download_report")
def download_report(video_url):
    """Generate and return Excel file for download"""
    print("=" * 60)
    print("📥 DOWNLOAD REPORT CALLED")
    print(f"Video URL: {video_url}")
    print("=" * 60)
    
    api_key = "AIzaSyCqrNNtoQyJQ7jOPvjUNTfHvC0MfdrEvYk"
    
    # Extract video ID
    video_id = None
    if "v=" in video_url:
        video_id = re.search(r"(?<=v=)[^&]+", video_url).group(0)
    elif "youtu.be/" in video_url:
        video_id = re.search(r"youtu\.be/([^?]+)", video_url).group(1)
    
    print(f"Extracted Video ID: {video_id}")
    
    if not video_id:
        print("❌ ERROR: Invalid video ID")
        return None
    
    try:
        # Fetch all data again
        print("📊 Fetching video metadata...")
        video_metadata = fetch_video_metadata(video_id, api_key)
        print(f"✓ Video: {video_metadata['title'] if video_metadata else 'Unknown'}")
        
        print("💬 Fetching comments...")
        comments = fetch_comments(video_id, api_key, max_comments=1000)
        print(f"✓ Fetched {len(comments)} comments")
        
        print("🔍 Analyzing comments...")
        analyzed_comments = analyze_all_comments(comments)
        print(f"✓ Analyzed {len(analyzed_comments)} comments")
        
        print("📈 Calculating overall sentiment...")
        overall_sentiment, overall_score = calculate_overall_sentiment(analyzed_comments)
        print(f"✓ Sentiment: {overall_sentiment} ({overall_score})")
        
        print("🎭 Classifying genre...")
        genre = classify_genre(analyzed_comments)
        print(f"✓ Genre: {genre}")
        
        # Create Excel file
        print("📄 Creating Excel file...")
        excel_buffer = create_excel_file(
            analyzed_comments, 
            video_metadata, 
            overall_sentiment, 
            overall_score, 
            genre
        )
        print(f"✓ Excel file created, size: {len(excel_buffer.getvalue())} bytes")
        
        # Create filename with video title and timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        video_title_clean = re.sub(r'[^\w\s-]', '', video_metadata['title'][:30]) if video_metadata else "video"
        filename = f"YouTube_Analysis_{video_title_clean}_{timestamp}.xlsx"
        print(f"✓ Filename: {filename}")
        
        # Return as Anvil media object
        print("📦 Creating Anvil media object...")
        # Get the bytes from BytesIO
        excel_bytes = excel_buffer.getvalue()
        print(f"✓ Got {len(excel_bytes)} bytes")
        
        # Create BlobMedia object
        media_obj = anvil.BlobMedia(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            excel_bytes,
            name=filename
        )
        print("✅ SUCCESS! Returning Excel file to Anvil")
        print("=" * 60)
        return media_obj
    
    except Exception as e:
        print(f"❌ ERROR generating report: {str(e)}")
        import traceback
        traceback.print_exc()
        print("=" * 60)
        return None

# ============================================================================
# START SERVER
# ============================================================================

print("=" * 60)
print("YouTube Video Sentiment Analyzer - Server Started")
print("=" * 60)
print("Waiting for requests from Anvil app...")
print("Features enabled:")
print("  ✓ VADER Sentiment Analysis")
print("  ✓ Per-comment Analysis")
print("  ✓ Excel Report Generation")
print("  ✓ Spam Detection")
print("  ✓ Bot Detection")
print("  ✓ Video Metadata")
print("  ✓ Enhanced Genre Classification")
print("=" * 60)

anvil.server.wait_forever()

